"""
oscodepoint.py
==============

An interface to Ordnance Survey's CodePoint-Open. CodePoint-Open is a free
dataset that maps UK postcodes to coordinates.

`oscodepoint` reads in this data, whether in the original zip or decompressed,
parses the data, and converts grid references to latitude and longitude.

The dataset can be downloaded from
http://www.ordnancesurvey.co.uk/oswebsite/products/code-point-open/


Example:
--------
    >>> from oscodepoint import open_codepoint
    >>> codepoint = open_codepoint('codepo_gb.zip')
    >>> for entry in codepoint.entries():
    ...    print entry['Postcode'], entry['Latitude'], entry['Longitude']
    ...    break  # Over 1.6 million rows
    AB101AA 57.1482995075 -2.09663094048


Too much data? Try limiting the postcode areas:
-----------------------------------------------
    >>> from oscodepoint import open_codepoint
    >>> codepoint = open_codepoint('codepo_gb.zip')
    >>> for entry in codepoint.entries(areas=['NR', 'IP']):
    ...    print entry['Postcode'], entry['Eastings'], entry['Northings']
    ...    break
    NR1 1AA 624068 308352


Want the postcode's county?
---------------------------
Postcode entries have a `Admin_county_code` field. `Doc/Codelist.xls` maps
these codes to county names, and `codepoint.codelist` can be used to access
this file. For example:

    >>> from oscodepoint import open_codepoint
    >>> codepoint = open_codepoint('codepo_gb.zip')
    >>> county_list = codepoint.codelist['County']
    >>> for entry in codepoint.entries(areas=['NR']):
    ...    print entry['Postcode'], entry['Latitude'], entry['Longitude'], county_list.get(entry['Admin_county_code'])
    ...    break
    NR1 1AA 52.6266175146 1.30932087485 Norfolk County


Get the total number of postcodes for your progress bar:
--------------------------------------------------------
    >>> from oscodepoint import open_codepoint
    >>> codepoint = open_codepoint('codepo_gb.zip')
    >>> print codepoint.metadata['area_counts']['NR']
    22730
    >>> print codepoint.metadata['total_count']
    1692241
"""


from collections import OrderedDict
import csv
import fnmatch
import glob
import os.path
import pyproj
import re
from StringIO import StringIO
import zipfile

import xlrd
import openpyxl

__all__ = ['open_codepoint', 'CodePointDir', 'CodePointZip']

class FileNotFoundError(Exception):
    pass

def open_codepoint(filename):
    """
    Open a CodePoint directory or zip file. Returns a CodePointDir or
    CodePointZip object.
    """

    if os.path.isdir(filename):
        return CodePointDir(filename)
    else:
        return CodePointZip(filename)


class lazyproperty(object):
    """
    Memoizing property. Calls `fget()` once, then stores the result.
    """

    def __init__(self, fget):
        self.fget = fget

    def __get__(self, obj, type=None):
        value = self.fget(obj)
        setattr(obj, self.fget.func_name, value)
        return value


class BaseCodePoint(object):
    """
    Abstract access to CodePoint data. You should use `CodePointZip`,
    `CodePointDir`, or just forget about the difference and use `open_codepoint()`.
    """

    root = ''
    headers_name = 'Doc/Code-Point_Open_Column_Headers.csv'
    metadata_name = 'Doc/metadata.txt'
    codelist_names = ['Doc/Codelist.xls', 'Doc/Codelist.xlsx']
    nhs_codelist_names = ['Doc/NHS_Codelist.xls', 'Doc/NHS_Codelist.xlsx']
    data_name_format = 'Data/CSV/%s.csv'

    def entries(self, areas=None, to_proj=pyproj.Proj(init='epsg:4326')):
        """
        Iterate over postcode entries.

        Limit the postcode areas with the `areas` parameter. Set to `None`
        (the default) to iterate over everything.

        Grid references are converted to latitude and longitude - the target
        coordinate system is defined by the `to_proj` parameter. Set it to a
        `pyproj.Proj` instance to change from the default of WGS84, or use
        `None` if you don't want coordinate conversion.
        """

        from_proj = pyproj.Proj(init='epsg:27700') # British National grid

        if areas is None:
            areas = self.areas

        for area in areas:
            if not re.search(r'^[A-Za-z]{1,2}$', area):
                raise ValueError('Incorrect format for area: '
                                 'expected 1 or 2 letters, got "%s"' % (area,))

            for row in self._get_name_rows(self.data_name_format % area.lower()):
                entry = OrderedDict(zip(self.long_headers, row))
                entry['_Area'] = area

                if to_proj is not None:
                    eastings, northings = float(entry['Eastings']), float(entry['Northings'])
                    lng, lat = pyproj.transform(from_proj, to_proj, eastings, northings)
                    entry['Longitude'], entry['Latitude'] = lng, lat

                yield entry

    @lazyproperty
    def areas(self):
        return list(self._get_areas())

    @lazyproperty
    def long_headers(self):
        return self._get_headers()['long']

    @lazyproperty
    def metadata(self):
        return self._get_metadata()

    @lazyproperty
    def codelist(self):
        return self._get_codelist()

    @lazyproperty
    def nhs_codelist(self):
        return self._get_nhs_codelist()

    def _areas_from_names(self, names):
        pattern = re.compile(r'[\\/]([a-z]{1,2})\.csv$')
        for name in names:
            match = pattern.search(name)
            if match:
                yield match.group(1)


    def _get_codelist(self):
        return self._construct_codelist(CodeList, self.codelist_names)

    def _get_nhs_codelist(self):
        return self._construct_codelist(NHSCodeList, self.nhs_codelist_names)



class CodePointZip(BaseCodePoint):
    """
    Read CodePoint data from a zip file.
    """

    def __init__(self, zip_filename):
        self.zip_file = zipfile.ZipFile(zip_filename)

    def _open(self, name):
        return self.zip_file.open(name)

    def _read(self, name):
        return self.zip_file.read(name)

    def _get_areas(self):
        pattern = self.data_name_format % '*'
        return self._areas_from_names(
            name for name in self.zip_file.namelist()
            if fnmatch.fnmatch(name, pattern)
        )

    def _get_name_rows(self, name):
        return csv.reader(self._open(name))

    def _get_headers(self):
        short_headers, long_headers = csv.reader(self._open(self.headers_name))
        return dict(
            short=short_headers,
            long=long_headers,
        )

    def _get_metadata(self):
        return Metadata(self._open(self.metadata_name))

    def _construct_codelist(self, list_class, potential_filenames):
        for filename in potential_filenames:
            try:
                contents = self._read(filename)
            except KeyError:
                continue # Try the next one in the list
            return list_class(filename, file_contents=contents)
        else:
            raise FileNotFoundError("Could not find code list file: tried %s" %
                                    (", ".join(potential_filenames)))


class CodePointDir(BaseCodePoint):
    """
    Read CodePoint data from a decompressed zip file.
    """

    def __init__(self, path):
        self.path = path
        if os.path.isdir(os.path.join(self.path, self.root)):
            self.path = os.path.join(self.path, self.root)

    def _get_areas(self):
        return self._areas_from_names(glob.glob(os.path.join(self.path, self.data_name_format % '*')))

    def _get_name_rows(self, name):
        return csv.reader(open(os.path.join(self.path, name)))

    def _get_headers(self):
        short_headers, long_headers = csv.reader(open(os.path.join(self.path, self.headers_name)))
        return dict(
            short=short_headers,
            long=long_headers,
        )

    def _get_metadata(self):
        return Metadata(open(os.path.join(self.path, self.metadata_name)))


    def _construct_codelist(self, list_class, potential_filenames):
        for filename in potential_filenames:
            full_pathname = os.path.join(self.path, filename)
            if os.path.exists(full_pathname):
                return list_class(full_pathname)
        else:
            raise FileNotFoundError("Could not find code list file: tried %s" %
                                    (", ".join(potential_filenames)))

    FOO = """
    def _get_codelist(self):
    def _get_nhs_codelist(self):
        return NHSCodeList(os.path.join(self.path, self.nhs_codelist_name))
    """

class Metadata(dict):
    """
    Parse the Doc/metadata.txt file. Used via `codepoint.metadata`
    """

    header_re = re.compile(r'^([^:]+):\s*([^:]+)$')
    area_count_re = re.compile(r'^\s+([A-Z]{1,2})\s+(\d+)$')

    def __init__(self, f):
        self['area_counts'] = {}
        for line, mode in self.line_modes(f):
            if mode == 'header':
                match = self.header_re.search(line)
                self[match.group(1)] = match.group(2)

            if mode == 'area_count':
                match = self.area_count_re.search(line)
                self['area_counts'][match.group(1)] = int(match.group(2))

        self['total_count'] = sum(self['area_counts'].itervalues())

    def line_modes(self, lines):
        mode = 'file_start'
        for line in lines:
            line = line.rstrip()
            mode = self.line_mode(line, mode)
            yield (line, mode)

    def line_mode(self, line, prev_mode):
        magic = 'ORDNANCE SURVEY'

        if prev_mode == 'file_start':
            if line == magic:
                return 'magic'
            else:
                raise ValueError('Expected "%s" text on first line of metadata file' % magic)

        if prev_mode in ('magic', 'header',):
            if self.header_re.search(line):
                return 'header'
            elif self.area_count_re.search(line):
                return 'area_count'

        if prev_mode == 'area_count':
            if self.area_count_re.search(line):
                return 'area_count'

        raise ValueError('Can\'t get next mode from mode "%s" and line "%s"' %
                         (prev_mode, line,))


class _ExcelCodeList(dict):
    """
    Base class for dicts generated from Excel files.
    """
    def __init__(self, filename, file_contents=None):
        if filename.lower().endswith(".xlsx"):
            book = self._load_xslx_workbook(filename, file_contents)
            self._init_from_xlsx(book)
        else:
            book = xlrd.open_workbook(filename, file_contents=file_contents)
            self._init_from_xls(book)

    def _load_xslx_workbook(self, filename, file_contents):
        if isinstance(file_contents, str):
            file_stream = StringIO(file_contents)
        else:
            file_stream = file_contents
        return openpyxl.load_workbook(file_stream or filename)

    def _get_mappings_from_xlsx_sheet(self, book, sheet_name):
        """Return a dict mapping codes to human-readable-names"""
        sheet = book[sheet_name]

        def _ordered_pair(first, second):
            if self.NAMES_BEFORE_CODES:
                return (second, first)
            else:
                return (first, second)

        return dict(
            _ordered_pair(first, second)
            for (first, second) in (
                (sheet.cell(row=row_index, column=sheet.min_column).value,
                 sheet.cell(row=row_index, column=sheet.max_column).value)
                for row_index in xrange(sheet.min_row, sheet.max_row+1)
            )
        )


class CodeList(_ExcelCodeList):
    """
    The CodePoint download has a Doc/Codelist.xls Excel-format spreadsheet.
    This has multiple worksheets, with one lookup table per sheet.
    `CodeList` reads in those lookup tables. Use it via `codepoint.codelist`.
    """

    NAMES_BEFORE_CODES = True

    def _populate_lookup_aliases(self, lookup_aliases):
        for alias, lookup_name in lookup_aliases.iteritems():
            self[alias] = self[lookup_name]

    def _init_from_xls(self, book):
        lookup_aliases = {}
        for sheet in book.sheets():
            if sheet.name == 'Metadata':
                # The metadata sheet doesn't have any lookups.
                continue

            self[sheet.name] = dict(
                (key, value)
                for (value, key) in (
                    sheet.row_values(row_index)
                    for row_index in xrange(sheet.nrows)
                )
            )

            if sheet.name == 'AREA_CODES':
                # The AREA_CODES sheet has a mapping of sheet names to
                # friendlier names. We'll use these at the end of the loop.
                lookup_aliases = self[sheet.name]

        self._populate_lookup_aliases(lookup_aliases)

    def _init_from_xlsx(self, book):
        lookup_aliases = {}
        for sheet_name in book.sheetnames:
            if sheet_name == 'Metadata':
                # The metadata sheet doesn't have any lookups.
                continue

            self[sheet_name] = self._get_mappings_from_xlsx_sheet(book,
                                                                  sheet_name)

            if sheet_name == 'AREA_CODES':
                # The AREA_CODES sheet has a mapping of sheet names to
                # friendlier names. We'll use these at the end of the loop.
                lookup_aliases = self[sheet_name]

        self._populate_lookup_aliases(lookup_aliases)


class NHSCodeList(_ExcelCodeList):
    """
    Similar to `CodeList`, but:
      * No Metadata or AREA_CODES worksheet.
      * The key and value columns are in the opposite order.
    """

    NAMES_BEFORE_CODES = False

    def _init_from_xls(self, book):
        for sheet in book.sheets():
            self[sheet.name] = dict(
                (key, value)
                for (key, value) in (
                    sheet.row_values(row_index)
                    for row_index in xrange(sheet.nrows)
                )
            )

    def _init_from_xlsx(self, book):
        for sheet_name in book.sheetnames:
            self[sheet_name] = self._get_mappings_from_xlsx_sheet(book,
                                                                  sheet_name)
